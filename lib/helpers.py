"Common helper functions"

from functools import lru_cache
from typing import Any, Callable, Iterable, Type, TypeVar

from openpyxl import Workbook, load_workbook

T = TypeVar("T")


class ExcelHelpers:
    """Helpers for Excel"""

    @classmethod
    def get_col_by_name(cls, sheet, name: str, header_row: int = 1) -> int:
        """
        Search row for a cell with a particular value and return it's column number
        Return value is 1-based
        """
        for col in sheet.iter_cols():
            for ind in range(
                1, 3
            ):  # checks previous row as well in case of merged cells
                cell = col[header_row - ind]
                if (
                    isinstance(cell.value, str)
                    and cell.value.strip().lower() == name.strip().lower()
                ):
                    return cell.column
        raise ValueError(f'Column "{name}" was not found')

    @classmethod
    def get_value_by_col_name(cls, sheet, name: str, row: list[str]) -> str:
        """Returns cell value of a row by the column name"""
        col_number = ExcelHelpers.get_col_by_name(sheet, name, 1) - 1
        return row[col_number]


class BaseWorkBook:
    """Base class for all classes which open Excel books"""

    workbook: Workbook
    filename: str

    def close(self):
        "Close the opened workbook"
        try:
            self.workbook.close()
        except (NameError, AttributeError):
            pass

    def save(self):
        "Save opneed workbook"
        self.workbook.save(self.filename)


class BaseWorkBookData(BaseWorkBook):
    "Base class representing data read from workbook"

    def get_row_by_field_value(self, field: str, value: str) -> Any:
        "Finds and returns first row where field == value"
        for record in self.records:
            if getattr(record, field) == value:
                return record
        raise ValueError(
            f"Field {field} with value {value} not found in file {self.filename}"
        )

    def as_filtered_list(self, fields: Iterable, values: Iterable) -> list[Any]:
        "Returns list of rows filtered by all values of a given pair of iterables"

        def _check_fields(record: Any, fields: Iterable, values: Iterable):
            for field, value in zip(fields, values):
                if getattr(record, field) != value:
                    return False
            return True

        return list(
            filter(lambda record: _check_fields(record, fields, values), self.records)
        )

    def get_field_values(self, field: str):
        "Returns sorted list of all possible values of a given field"
        return list(sorted({getattr(r, field) for r in self.records}))

    def __init__(
        self,
        filename: str,
        header_row: int,
        record_class: Type[T],
        filter_func: Callable[[Any], bool] | None = None,
        max_col: int | None = None,
    ) -> None:
        self.filename = filename
        self.records: list[T] = list()
        try:
            self.workbook = load_workbook(filename=filename, data_only=True)
            sheet = self.workbook.active
            assert sheet is not None
            for row in sheet.iter_rows(  # type: ignore
                min_row=header_row + 1, max_col=max_col, values_only=True
            ):
                record: T = record_class(*row)
                if callable(filter_func) and not filter_func(record):
                    continue
                self.records.append(record)
        finally:
            try:
                self.close()
            except AttributeError:
                pass


class BaseMultisheetWorkBookData(BaseWorkBook):
    "Base class representing multiple data sheets read from workbook"

    def get_sheets_count(self) -> int:
        "Returns total number of sheets read from file"
        return len(self.sheets)

    def get_strings_count(self) -> int:
        "Returns total number of data strings of all sheets read from file"
        return sum([len(s) for _, s in self.sheets.items()])

    def get_row_by_field_value(self, field: str, value: str, sheet_name: str) -> Any:
        "Finds in a given list and returns first row where field == value"
        for record in self.sheets[sheet_name]:
            if getattr(record, field) == value:
                return record
        raise ValueError(
            f"Field {field} with value {value} not found in a sheet {sheet_name} "
            "of a file {self.filename}"
        )

    @lru_cache(maxsize=1)
    def get_account_row(self, account: str, sheet_name: str) -> Any:
        "Finds on a given sheet and returns a row with given value of .account attribute"
        return self.get_row_by_field_value("account", account, sheet_name)

    def as_filtered_list(
        self, fields: Iterable, values: Iterable, sheet_name: str
    ) -> list[Any]:
        "Returns list of rows filtered by all values of a given pair of iterables"

        def _check_fields(record: Any, fields: Iterable, values: Iterable):
            for field, value in zip(fields, values):
                if getattr(record, field) != value:
                    return False
            return True

        return list(
            filter(
                lambda record: _check_fields(record, fields, values),
                self.sheets[sheet_name],
            )
        )

    def __init__(
        self,
        filename: str,
        header_row: int,
        record_class: Type[T],
        filter_func: Callable[[Any], bool] | None = None,
        max_col: int | None = None,
    ) -> None:
        self.filename = filename
        self.sheets: dict[str, list[T]] = dict()
        try:
            self.workbook = load_workbook(filename=filename, data_only=True)
            for sheet in self.workbook:
                records = []
                for row in sheet.iter_rows(
                    min_row=header_row + 1, max_col=max_col, values_only=True
                ):
                    record: T = record_class(*row)
                    if callable(filter_func) and not filter_func(record):
                        continue
                    records.append(record)
                self.sheets[sheet.title] = records
        finally:
            try:
                self.close()
            except AttributeError:
                pass


class SingletonWithArg(type):
    """
    Singleton hashed by the __init__() first string argument,
    """

    _instances: dict = {}

    def __call__(cls, arg: str, *args, **kwargs):
        key = (cls, arg)
        if key not in cls._instances:
            cls._instances[key] = super(SingletonWithArg, cls).__call__(
                arg, *args, **kwargs
            )
        return cls._instances[key]
