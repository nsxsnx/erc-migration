"Base class to work with adress-formatted Excel tables"

from dataclasses import dataclass
import logging
import string

from openpyxl import load_workbook

from lib.helpers import BaseWorkBook, ExcelHelpers

SheetDataFormatted = list[str]
SheetDataRaw = list[tuple[str]]


@dataclass
class BuildingRecord:
    "Record of Buildings table"
    num: int
    municipality: str
    street: str
    house: str
    building: str
    correction_month: int


class AddressFile(BaseWorkBook):
    """Represents adress list"""

    def get_sheet_data_formatted(self, name: str) -> list[str]:
        "Returns valuable strings of a sheet (formatted to string)"
        if name in self._data_formatted:
            return self._data_formatted[name]
        raise KeyError(f'Sheet "{name}" not found in "{self.filename}"')

    def get_sheet_data_raw(self, name: str) -> list[list]:
        "Returns valuable strings of a sheet (raw)"
        if name in self._data_raw:
            return self._data_raw[name]
        raise KeyError(f'Sheet "{name}" not found in "{self.filename}"')

    def _get_sheet_data(self, sheet) -> tuple[SheetDataFormatted, SheetDataRaw]:
        formatted, raw = list(), list()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            fields: dict = {}
            for field_name in self.fields:
                if not field_name:
                    raise ValueError(
                        "Attribute field can't be empty, check ADDRESS_FORMAT in config"
                    )
                field_value = ExcelHelpers.get_value_by_col_name(sheet, field_name, row)
                fields.update({field_name: field_value})
            formatted.append(self.record_format.format(**fields).lower())
            raw.append(row)
        if not formatted:
            logging.warning("Sheet %s seems empty!", sheet.title)
        return formatted, raw

    def get_row_by_address(self, name: str, address: str) -> tuple[str]:
        "Finds and returns row data for a given address in a giben sheet"
        sheet_addresses = [f"{a}," for a in self.get_sheet_data_formatted(name)]
        address = address.lower()
        for counter, sheet_address in enumerate(sheet_addresses):
            if sheet_address in address:
                return self.get_sheet_data_raw(name)[counter]
        raise ValueError(
            f"Address '{address}' not found in sheet '{name}' of '{self.filename}'"
        )

    def __init__(self, file_name: str, conf: dict) -> None:
        self.filename = file_name
        self.record_format: str = conf["address_format.street"]
        self.fields = [v[1] for v in string.Formatter().parse(self.record_format)]
        self._data_formatted: dict = {}
        self._data_raw: dict = {}
        try:
            self.workbook = load_workbook(filename=self.filename, data_only=True)
            for sheet in self.workbook:
                data_formatted, data_raw = self._get_sheet_data(sheet)
                self._data_formatted.update({sheet.title: data_formatted})
                self._data_raw.update({sheet.title: data_raw})
        finally:
            self.close()
        logging.debug(self._data_formatted)
        super().__init__()

    def get_sheets_count(self) -> int:
        "Returns total number of sheets read from file"
        return len(self._data_formatted)

    def get_strings_count(self) -> int:
        "Returns total number of data strings of all sheets read from file"
        return sum([len(s) for _, s in self._data_formatted.items()])
