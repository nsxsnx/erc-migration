"Base class to work with OSV-formatted Excel tables"

from decimal import Decimal
from functools import total_ordering
import logging
from pathlib import Path
import re
import sys
from dataclasses import dataclass
from os.path import basename
from typing import Generator, Self

from openpyxl import load_workbook
from lib.buildingsfile import BuildingsFile

from lib.datatypes import MonthYear
from lib.exceptions import NoAddressRow
from lib.helpers import BaseWorkBook, ExcelHelpers

OSVDATA_REGEXP = [
    r"^(?P<type>Частная|Муниципальная|Служебная|Общежитие|Частная без регистр.|"
    r"Собственн.юридич.лиц|Арендуемая|Маневренный фонд|Приватизированная|),"
    r"(?P<account>\d{12}),"
    r"((?P<name>.*);)?"
    r" ?(?P<address>(ул|мкр) .*),"
    r"чел.-(?P<population>[\d\.]{1,2}) "
    r"площ.-(?P<area>[\d\.]{1,9}),"
    r"(?P<status>Открыт|Закрыт|Пустующий)$",
]

osvdata_regexp_compiled: list[re.Pattern] = []


@dataclass
class OsvAccuralRecord:
    'Stores OSV "accural" values'
    heating: Decimal
    gvs: Decimal
    reaccural: Decimal
    payment: Decimal
    gvs_elevated_percent: Decimal


@dataclass
class OsvAddressRecord:
    'Stores OSV "address" column data'
    type: str
    account: str
    name: str
    address: str
    population: str
    area: str
    status: str

    @classmethod
    def get_instance(cls, data: str) -> Self:
        "Returns new instalnce of itself"
        for expr in osvdata_regexp_compiled:
            if match := re.match(expr, data):
                return cls(**match.groupdict())
            else:
                continue
        raise ValueError(f"Can't understand value: {data}")


@dataclass
class OsvRecord:
    "Stores OSV row"
    address: OsvAddressRecord
    accural: OsvAccuralRecord


@dataclass
class OsvColumnIndex:
    "Indexes of fields in the table, zero-based"
    address: int
    heating: int
    gvs: int
    reaccurance: int
    total: int
    gvs_elevated_percent: int

    @classmethod
    def from_workbook(cls, workbook, header_row, headers) -> Self:
        "Calculates indexes of columns in the table"

        if workbook is None:
            raise ValueError("OSV file was not initialized yet")
        try:
            return cls(
                *[
                    ExcelHelpers.get_col_by_name(workbook.sheet, header, header_row) - 1
                    for header in headers
                ]
            )
        except ValueError as err:
            logging.warning("Check OSV column names and quantity: %s", err)
            raise


class OsvFile(BaseWorkBook):
    """Represents an OSV-file"""

    def _init_date(self) -> None:
        """Reads OSV date from file"""
        cell_value = self.sheet[self.conf["osv.date_cell"]].value  # type: ignore
        date_match = re.match(r"^[\s\S]* (\d{1,2})\.(\d{4})$", cell_value)
        if not date_match:
            raise ValueError(f"Date not found in OSV file header: {self.filename}")
        self.date = MonthYear(int(date_match.group(1)), int(date_match.group(2)))
        if not self.date.month or not self.date.year:
            raise ValueError(f"Incorrect date in OSV file: {self.filename}")
        logging.debug("OSV date: %d.%d", self.date.month, self.date.year)

    def _init_column_indexes(self) -> None:
        self.column_indexes = OsvColumnIndex.from_workbook(
            self,
            int(self.conf["osv.header_row"]),
            [
                "Адрес",
                "Отопление",
                "Тепловая энергия для подогрева воды",
                "Перерасчеты",
                "Всего",
                "Тепловая энергия для подогрева воды (повышенный %)",
            ],
        )

    def __init__(self, file: str, conf: dict) -> None:
        self.conf = conf
        logging.info("Reading OSV: %s...", basename(file))
        self.filename = file
        self.workbook = load_workbook(filename=file, data_only=True)
        self.sheet = self.workbook.active
        self.record: OsvRecord
        try:
            self._init_date()
        except (AttributeError, NameError):
            logging.critical("Can't get date from OSV file")
            self.close()
            sys.exit(1)
        self._init_column_indexes()
        super().__init__()

    def _get_row(self):
        """Generator that reads OSV-data line by line"""
        if self.sheet is None:
            raise StopIteration
        for row in self.sheet.iter_rows(  # type:ignore
            min_row=int(self.conf["osv.header_row"]) + 1, values_only=True
        ):
            if row[1]:
                yield row
            else:
                continue

    def close(self):
        self.workbook.close()

    def init_next_record(
        self, buildings: BuildingsFile
    ) -> Generator[OsvRecord, None, None]:
        "Parses OSV row and returns OvsRecord instance"

        for row in self._get_row():
            address_cell = row[self.column_indexes.address]
            try:
                osv_address_rec = OsvAddressRecord.get_instance(address_cell)
                logging.debug(
                    "Address record %s understood as %s", row[0], osv_address_rec
                )
                try:
                    buildings.get_address_row(
                        osv_address_rec.address, str(self.date.year)
                    )
                except NoAddressRow:
                    continue
                osv_accural_rec = OsvAccuralRecord(
                    Decimal(row[self.column_indexes.heating]),
                    Decimal(row[self.column_indexes.gvs]),
                    Decimal(row[self.column_indexes.reaccurance]),
                    Decimal(row[self.column_indexes.total]),
                    Decimal(row[self.column_indexes.gvs_elevated_percent]),
                )
                logging.debug(
                    "Accural record %s understood as %s", row[0], osv_accural_rec
                )
            except AttributeError as error:
                logging.warning("%s. Malformed record: %s", error, row)
                continue
            self.record = OsvRecord(osv_address_rec, osv_accural_rec)
            yield self.record


@total_ordering
class OsvPath(type(Path())):
    "Path with custom sorting behavior and validation"

    def __new__(cls, *args, **kwargs) -> Self:
        return super().__new__(cls, *args, **kwargs)

    def __lt__(self, other):
        self_date = MonthYear(int(self.name[0:2]), int(self.name[2:6]))
        other_date = MonthYear(int(other.name[0:2]), int(other.name[2:6]))
        return self_date < other_date

    def validate(self):
        "Checks if ``Self`` conforms to some requirements"

        if not self.with_suffix(".xlsx"):
            logging.critical("Non *.xlsx found in OSV_DIR, exiting")
            sys.exit(1)
